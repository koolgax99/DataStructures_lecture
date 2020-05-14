import xlrd
import openpyxl

def donordata():
    print("\nEnter the following details:")
    name = input("Fund donor's name : ")
    amtdon = input("Amount donated : ")
    waydon = input("Enter the method of donation : ")
    waydon.lower()

    file = 'data.xlsx'
    new_row = [name, amtdon, waydon]
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet1']   
    row = ws.max_row + 1
    for col, entry in enumerate(new_row, start=1):
        ws.cell(row=row, column=col, value=entry)
    wb.save(file)
            
    print("\nThank you! Fund donor's details successfully updated!")

def recipientdata():
    print("\nEnter the following details:")
    name = input("Fund recipient's name : ")
    amtdon = input("Amount donated : ")
    waydon = input("Enter the method of donation : ")
    waydon.lower()

    file = 'data.xlsx'
    new_row = [name, amtdon, waydon]
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Sheet2']    
    row = ws.max_row + 1
    for col, entry in enumerate(new_row, start=1):
        ws.cell(row=row, column=col, value=entry)
    wb.save(file)
            
    print("\nThank you! Fund recipient's details successfully updated!")

def retrievedonordata():
                wb1 = xlrd.open_workbook("data.xlsx") 
                sheet = wb1.sheet_by_index(0)
                sheet.cell_value(0, 0)
                x = sheet.nrows


                print("\n1. Show data of all fund donors\n2. Show data of specific method of fund transfer\n")
                opt = int(input("Enter choice : "))
                if opt==1: 
                    
                    for i in range(x): 
                        print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)) +'  '+ sheet.cell_value(i,2))

                    

                elif opt==2:
                    hosp = input("Enter fund transfer method : ")
                    hosp.lower()
                    if(hosp == "cash"):
                        for i in range(sheet.nrows): 
                            if(sheet.cell_value(i, 2)=="cash"):
                                print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)))           
                    

                    elif(hosp == "cheque"):
                        for i in range(sheet.nrows): 
                            if(sheet.cell_value(i, 2)=="cheque"):
                                print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)))

                    else: {
                        print("Invalid Method!!!")
                    }

                else:
                    print("Wrong choice!! Try again!")

def retrieverecipientdata():
                wb1 = xlrd.open_workbook("data.xlsx") 
                sheet = wb1.sheet_by_index(1)
                sheet.cell_value(0, 0)
                x = sheet.nrows


                print("\n1. Show data of all fund recipients\n2. Show data of specific method of fund transfer\n")
                opt = int(input("Enter choice : "))
                if opt==1: 
                    
                    for i in range(x): 
                        print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)) +'  '+ sheet.cell_value(i,2))

                    

                elif opt==2:
                    hosp = input("Enter fund transfer method : ")
                    hosp.lower()
                    if(hosp == "cash"):
                        for i in range(sheet.nrows): 
                            if(sheet.cell_value(i, 2)=="cash"):
                                print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)))           
                    

                    elif(hosp == "cheque"):
                        for i in range(sheet.nrows): 
                            if(sheet.cell_value(i, 2)=="cheque"):
                                print(sheet.cell_value(i, 0) +'  '+str(sheet.cell_value(i,1)))

                    else: {
                        print("Invalid Method!!!")
                    }

                else:
                    print("Wrong choice!! Try again!")

def exitportal():
    print("\n\nThank you for using our system!!\nHave a good day!\n")
    sys.exit(1)



def main():
    while True:
        print("\n\n1. Enter fund donor's data\n2. Enter fund recipient's data\n3. Retrieve Donor specific data\n4. Retrieve Recipient specific data\n5. Exit system\n\n")
        ch = int(input("Enter your choice : "))
        if ch==1:{
            donordata()
        }
        elif(ch==2): {
            recipientdata()    
        }
        elif(ch==3): {
            retrievedonordata()
        }
        elif(ch==4): {
            retrievereciepientdata()
        }
        elif(ch==5): {
            exitportal()
        }
        else:
            print("Wrong Choice!! Try again!")


main()